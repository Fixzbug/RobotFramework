import openpyxl
import os

# Define the predefined title
predefined_title = "example_title"

# Function to append data to the Excel file at a specific sheet


def save_data(predefined_title, new_data, filename="data.xlsx", sheet_name="Sheet1"):
    # Check if the file exists
    file_exists = os.path.isfile(filename)

    if file_exists:
        # Load the existing workbook and check if the sheet exists
        workbook = openpyxl.load_workbook(filename)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
    else:
        # Create a new workbook and add the sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        # Add title to the first row for new workbook
        sheet.append([predefined_title])

    # Find the column index for the predefined title
    column_index = None
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == predefined_title:
            column_index = col[0].column
            break

    # If the title does not exist, add a new column
    if column_index is None:
        column_index = sheet.max_column + 1
        sheet.cell(row=1, column=column_index, value=predefined_title)

    # Append the data to the next available row in the identified column
    # Start at row 2 because row 1 is the header
    for i, data in enumerate(new_data, start=2):
        sheet.cell(row=i, column=column_index, value=data)

    # Save the workbook
    workbook.save(filename)


filename = "D:/automate/Excel_sync_robot_file/data_col.xlsx"

sheet_name = "DATA"

# Define the predefined title
predefined_title = "example_title"

# Example usage
new_data = ["data1", "data2", "data3"]
save_data(predefined_title, new_data, filename, sheet_name)
