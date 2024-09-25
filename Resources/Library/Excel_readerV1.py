import pandas  # 2.1.0
import json
import pandas as pd
import openpyxl  # 3.1.2
from openpyxl import load_workbook

import warnings
import win32com.client as win32  # 306
import random
import psutil


from openpyxl.utils.exceptions import InvalidFileException
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# - update 05/07/2024 by mach change old script
# - update 12/07/2024 by mach change def write
# - update 15/07/2024 by mach change new def write


def Read_Excel_by_Sheet(filepath, sheet_name):

    # เปิดไฟล์ Excel โดยใช้ openpyxl
    workbook = load_workbook(filename=filepath, read_only=True)
    # ตรวจสอบว่า sheet_name ที่ระบุมีอยู่ในไฟล์หรือไม่
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the file.")

    # อ่านข้อมูลจาก Excel ด้วย pandas
    excel_data_df = pd.read_excel(filepath, sheet_name=sheet_name, dtype=str)

    # ปิดไฟล์ Excel
    workbook.close()

    # ใช้เมธอด strip เพื่อลบช่องว่างด้านหน้าและด้านหลังในแต่ละเซลล์
    excel_data_df = excel_data_df.apply(lambda col: col.map(
        lambda x: x.strip() if isinstance(x, str) else x))

    # แปลง DataFrame ที่ทำความสะอาดแล้วเป็น JSON
    json_str = excel_data_df.to_json(orient='records')
    json_obj = json.loads(json_str)

    # นับจำนวนแถว
    row_count = len(excel_data_df.index)
    # นับจำนวนคอลัมน์
    column_count = len(excel_data_df.columns)

    # แสดงผลข้อมูล จำนวนแถว และจำนวนคอลัมน์
    print("Data:", json_obj, "ROW:", row_count, "COL:", column_count)
    return json_obj, row_count, column_count


def Write_Excel_by_Sheet(filepath, sheet_name, row_expect, column_expect, data):
    try:
        # Initialize the Excel application
        excel = win32.Dispatch("Excel.Application")

        # ทำให้ Excel ไม่ปรากฎบนหน้าจอ (ทำงานในพื้นหลัง)
        excel.Visible = False

        # อ้างอิงไปยัง workbook ที่เปิดอยู่ในปัจจุบัน
        workbook = excel.Workbooks.Open(filepath)

        # ตรวจสอบว่ามี sheet ที่ต้องการใน workbook หรือไม่
        try:
            sheet = workbook.Sheets(sheet_name)
        except Exception as e:
            print(f"Sheet '{sheet_name}' ไม่พบใน workbook.")
            workbook.Close(SaveChanges=False)
            return

        # ตัดช่องว่างที่ข้างหน้าและข้างหลังของข้อมูล
        trimmed_data = str(data).strip()

        # เขียนข้อมูลลงในเซลล์ที่ระบุ
        sheet.Cells(row_expect, column_expect).Value = trimmed_data

        # บันทึก workbook
        workbook.Save()

        # ปิด workbook และ Excel
        workbook.Close(SaveChanges=True)
        excel.Quit()

        # Terminate Excel process to ensure it's closed
        for proc in psutil.process_iter():
            if "EXCEL.EXE" in proc.name():
                proc.kill()

        # Print success message
        print("Data written successfully to", filepath, "in sheet", sheet_name,
              "starting at row", row_expect, "and column", column_expect, "data", data)

    except Exception as e:
        print(f"An error occurred: {str(e)}")


if __name__ == "__main__":
    # Example usage
    Read_Excel_by_Sheet(
        "D:\Automate\RobotFramework\Excel_sync_robot_file\Data_for_ClosingContract_All.xlsx", "Request No")
    # Write_Excel_by_Sheet("D:\Automate\RobotFramework\Excel_sync_robot_file\Data_for_ClosingContract_All.xlsx","Request No",2,2,str(random.randint(1,100)))
