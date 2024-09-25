import openpyxl
import os


# Function to append data horizontally to the Excel file at a specific sheet
def save_data(predefined_title, new_data, filename, sheet_name):
    # Check if the file exists
    file_exists = os.path.isfile(filename)

    if file_exists:
        # Load the existing workbook and check if the sheet exists
        workbook = openpyxl.load_workbook(filename)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
    else:
        # Create a new workbook and add the sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

    # Find the row index for the predefined title
    row_index = None
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        if row[0].value == predefined_title:
            row_index = row[0].row
            break

    # If the title does not exist, add a new row
    if row_index is None:
        row_index = sheet.max_row + 1
        sheet.cell(row=row_index, column=1, value=predefined_title)

    # Append the data horizontally to the identified row
    # Start at column 2 because column 1 is the title
    for i, data in enumerate(new_data, start=2):
        sheet.cell(row=row_index, column=i, value=data)

    # Save the workbook
    workbook.save(filename)


filename = "D:/automate/Excel_sync_robot_file/data_row.xlsx"

sheet_name = "DATA"

# Define the predefined title
predefined_title = "example_title"

# Example usage
new_data = ["data1", "data2", "data3"]
save_data(predefined_title, new_data, filename, sheet_name)
